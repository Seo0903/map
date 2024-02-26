import openpyxl

# 1. 액셀 파일 열기
map_name = "공통점"
workbook = openpyxl.load_workbook(f'E:/mapList/{map_name}.xlsx')
worksheet = workbook['Sheet1']
ws = workbook.active

# 데이터를 저장할 이중 리스트 생성

person = [] #출제자 G
point = [] #중요포인트 H
singcom = [] #노래묶음 I
sing1 =[] #노래1 J
sing2 =[] #노래2 K
sing3 =[] #노래3 L
sing4 =[] #노래4 M
fq1 =[] #앞지문1 N
bq1 =[] #뒷지문1 O
fq2 =[] #앞지문2 P
bq2 =[] #뒷지문2 Q
fq3 =[] #앞지문3 R
bq3 =[] #뒷지문3 S
ans1t = []
ans2t = []
ans3t = []

# G열부터 S열까지 데이터 읽어와서 각 리스트에 추가
for row in range(2, worksheet.max_row + 1):  # 2번째 행부터 시작
    person.append([worksheet.cell(row=row, column=7).value])  # G열 데이터 추가
    point.append([worksheet.cell(row=row, column=8).value])  # H열 데이터 추가
    singcom.append([worksheet.cell(row=row, column=9).value])  # I열 데이터 추가
    sing1.append([worksheet.cell(row=row, column=10).value])  # J열 데이터 추가
    sing2.append([worksheet.cell(row=row, column=11).value])  # K열 데이터 추가
    sing3.append([worksheet.cell(row=row, column=12).value])  # L열 데이터 추가
    sing4.append([worksheet.cell(row=row, column=13).value])  # M열 데이터 추가
    fq1.append([worksheet.cell(row=row, column=14).value])  # N열 데이터 추가
    bq1.append([worksheet.cell(row=row, column=15).value])  # O열 데이터 추가
    fq2.append([worksheet.cell(row=row, column=16).value])  # P열 데이터 추가
    bq2.append([worksheet.cell(row=row, column=17).value])  # Q열 데이터 추가
    fq3.append([worksheet.cell(row=row, column=18).value])  # R열 데이터 추가
    bq3.append([worksheet.cell(row=row, column=19).value])  # S열 데이터 추가
    ans1t.append([worksheet.cell(row=row, column=20).value.split('/')[0]])
    ans2t.append([worksheet.cell(row=row, column=21).value.split('/')[0]])
    ans3t.append([worksheet.cell(row=row, column=22).value.split('/')[0]])

ans1 = [] #정답1 T 20
ans2 = [] #정답2 U 21
ans3 = [] #정답3 V 22

anss = []

for row in range(2, worksheet.max_row + 1):
    ans1.append(worksheet.cell(row=row, column=20).value.split('/'))
    ans2.append(worksheet.cell(row=row, column=21).value.split('/'))
    ans3.append(worksheet.cell(row=row, column=22).value.split('/'))


for i in range(len(ans1)):
    anss.append(ans1[i]+ans2[i]+ans3[i]) 
    
# 각 문자열이 처음으로 나타나는 위치를 기억하기 위한 딕셔너리
first_occurrence = {}

# 문자열에 일련번호를 부여할 리스트
numbering = []

# 문자열을 순회하면서 처음으로 나타나는 위치를 기억
for idx, sublist in enumerate(anss):
    for data in sublist:
        if data not in first_occurrence:
            first_occurrence[data] = idx + 1  # 1부터 시작하도록 인덱스에 1을 더해줌

# 각 데이터에 일련번호를 부여
for sublist in anss:
    sublist_numbering = []
    for data in sublist:
        number = first_occurrence[data]
        sublist_numbering.append(number)
    numbering.append(sublist_numbering)
    
for i in range(len(person)):
    person[i][0] = 'Db("' + str(person[i][0]) + '")'
    point[i][0] = 'Db("' + str(point[i][0]) + '")'
    singcom[i][0] = 'Db("' + str(singcom[i][0]) + '")'
    sing1[i][0] = 'Db("' + str(sing1[i][0]) + '")'
    sing2[i][0] = 'Db("' + str(sing2[i][0]) + '")'
    sing3[i][0] = 'Db("' + str(sing3[i][0]) + '")'
    sing4[i][0] = 'Db("' + str(sing4[i][0]) + '")'
    fq1[i][0] = 'Db("' + str(fq1[i][0]) + '")'
    bq1[i][0] = 'Db("' + str(bq1[i][0]) + '")'
    fq2[i][0] = 'Db("' + str(fq2[i][0]) + '")'
    bq2[i][0] = 'Db("' + str(bq2[i][0]) + '")'
    fq3[i][0] = 'Db("' + str(fq3[i][0]) + '")'
    bq3[i][0] = 'Db("' + str(bq3[i][0]) + '")'
    ans1t[i][0] = 'Db("' + str(ans1t[i][0]) + '")'
    ans2t[i][0] = 'Db("' + str(ans2t[i][0]) + '")'
    ans3t[i][0] = 'Db("' + str(ans3t[i][0]) + '")'

unique_anss = list(set(anss))
    
ans = []

for i in range(len(unique_anss)):
    ans.append(f"{unique_anss}: {numbering+1}")

with open('output.txt', 'w') as f:
    for item in ans:
        f.write("%s\n" % item)
        f.write('\n')
        f.write('\n')
    f.write('const person = \n[0,')
    for row in person:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const point = \n[0,')
    for row in point:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const singcom = \n[0,')
    for row in singcom:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const sing1 = \n[0,')
    for row in sing1:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const sing2 = \n[0,')
    for row in sing2:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const sing3 = \n[0,')
    for row in sing3:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const sing4 = \n[0,')
    for row in sing4:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const fq1 = \n[0,')
    for row in fq1:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const bq1 = \n[0,')
    for row in bq1:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const fq2 = \n[0,')
    for row in fq2:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const bq2 = \n[0,')
    for row in bq2:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const fq3 = \n[0,')
    for row in fq3:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const bq3 = \n[0,')
    for row in bq3:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const ans1t = \n[0,')
    for row in ans1t:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const ans2t = \n[0,')
    for row in ans2t:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    f.write('const ans3t = \n[0,')
    for row in ans3t:
        f.write('[' + ', '.join(row) + '],\n')
    f.write('];\n')
    f.write('\n')
    
    